import xlwt
import openpyxl
import os
import datetime
import shutil
import json
import csv
from dateutil import tz

from app.myglobals import logfolder, appfolder 

from app.models.mysql import db_mysql

def gen_csv_by_query(myquery, filename):
    # 1.prepare table heads
    heads_page = [
        'id', #1
        'device_type', #2
        'vendor', #3
        'firmware', #4
        'rssi_ble', #5
        'rssi_wifi', #6
        'mac_ble', #7
        'mac_wifi', #8
        'mcu', #9
        'result_intensity_check', #10
        'result_commands_check', #11
        'result_broadcast_check', #12
        'result_type_check', #13
        'result_version_check', #14
        'result_mac_check', #15
        'datetime', #16
    ]
    heads_db = [
        'id', #1
        'devicecode', #2
        'factorycode', #3
        'fw_version', #4
        'rssi_ble1', #5
        'rssi_wifi1', #6
        'mac_ble', #7
        'mac_wifi', #8
        'reserve_str_1', #9
        'bool_qualified_signal', #10
        'bool_qualified_check', #11.1
        'status_cmd_check1', #11.2
        'bool_qualified_scan', #12
        'bool_qualified_deviceid', #13
        'reserve_bool_1', #14
        'reserve_int_1', #15
        'datetime', #16
    ]
    # 2.prepare table data
    # datas = tableclass.query.all()
    datas = myquery.all()
    # 3.prepare csv writer
    f = open(filename, 'w')
    csv_writer = csv.writer(f)
    # 4. insert table head row
    csv_writer.writerow(heads_page)
    # 5.insert data rows
    for data in datas:
        data_dict = data.__dict__
        data_dict.update({'datetime': data_dict.get('datetime').strftime('%Y-%m-%d %H:%M:%S')})
        # data_dict.pop('_sa_instance_state')
        # data_dict.pop('rssi_ble2')
        # data_dict.pop('rssi_wifi2')
        # data_dict.pop('status_cmd_check2')
        # data_dict.pop('bool_uploaded')
        # data_dict.pop('bool_qualified_overall')
        # data_list = list(data_dict.values())
        data_list = list()
        for item in heads_db:
            data_item = data_dict.get(item)
            if item == 'devicecode':
                data_item = data.device.name
            if item == 'factorycode':
                data_item = data.factory.name
            if item == 'reserve_int_1':
                data_item = True if data_item == 1 else False
            if item == 'bool_qualified_check':
                data_bool_qualified_check = data_item
                continue
            if item == 'status_cmd_check1':
                data_status_cmd_check1 = data_item
                data_item = "{0}({1})".format(data_bool_qualified_check, data_status_cmd_check1)
            data_list.append(data_item)
        csv_writer.writerow(data_list)
    # 6.save
    f.close()


def gen_csv_by_tableclass_query_legacy(tableclass, myquery, filename):
    # 1.prepare table heads
    tablename = tableclass.__tablename__
    heads_raw = db_mysql.metadata.tables.get(tablename).c
    heads = list()
    for item in heads_raw:
        heads.append(str(item).replace(tablename+'.','',1))
    # 2.prepare table data
    # datas = tableclass.query.all()
    datas = myquery.all()
    # 3.prepare csv writer
    f = open(filename, 'w')
    csv_writer = csv.writer(f)
    # 4. insert table head row
    csv_writer.writerow(heads)
    # 5.insert data rows
    for data in datas:
        data_dict = data.__dict__
        data_dict.pop('_sa_instance_state')
        data_dict.update({'datetime': data_dict.get('datetime').strftime('%Y-%m-%d %H:%M:%S')})
        data_list = list(data_dict.values())
        # todo: sort data_list
        csv_writer.writerow(data_list)
    # 6.save
    f.close()

def gen_xls_by_tableclass_query_legacy(tableclass, myquery, filename):
    # 1.prepare table heads
    tablename = tableclass.__tablename__
    heads_raw = db_mysql.metadata.tables.get(tablename).c
    heads = list()
    for item in heads_raw:
        heads.append(str(item).replace(tablename+'.','',1))
    # 2.prepare table data
    # datas = tableclass.query.all()
    datas = myquery.all()
    # 3.prepare excel object
    book = openpyxl.Workbook()
    sheet1 = book.create_sheet(index=0, title='sheet1')
    # 4. insert table head row
    # diff with xlwt, openpyxl row start at least 1
    row = 1
    for col,field in enumerate(heads):
        # sheet1.write(0, col, field)
        sheet1.cell(row, col+1).value = field
    # 5.insert data rows
    row += 1
    for data in datas:
        col = 0
        while col < len(heads):
            cell = data.__dict__.get(heads[col])
            if heads[col] == 'datetime':
                cell = cell.strftime('%Y-%m-%d %H:%M:%S')
            # sheet1.write(row, col, cell)
            # diff with xlwt, openpyxl col start at least 1
            sheet1.cell(row, col+1).value = cell
            col += 1
        row += 1
    # 6.save
    book.save(filename)

def gen_excel_legacy(tableclass, filename):
    # 1.prepare table heads
    # heads_raw = db_mysql.metadata.tables.get('testdatas').c
    tablename = tableclass.__tablename__
    heads_raw = db_mysql.metadata.tables.get(tablename).c
    heads = list()
    for item in heads_raw:
        # heads.append(str(item).replace('testdatas.','',1))
        heads.append(str(item).replace(tablename+'.','',1))
    # 2.prepare table data
    # datas = Testdata.query.all()
    datas = tableclass.query.all()
    # 3.prepare excel object
    book = xlwt.Workbook()
    sheet1 = book.add_sheet('sheet1')
    # todo
    # dateFormat = xlwt.XFStyle()
    # dateFormat.num_format_str = '%y-%m-%d %h:%m:%s'
    # dateFormat.num_format_str = 'yyyy/mm/dd'
    # 4. insert table head row
    for col,field in enumerate(heads):
        sheet1.write(0, col, field)
    # 5.insert data rows
    row = 1
    for data in datas:
        col = 0
        while col < len(heads):
            cell = data.__dict__.get(heads[col])
            if heads[col] == 'datetime':
                cell = cell.strftime('%Y-%m-%d %H:%M:%S')
            sheet1.write(row, col, cell)
            col += 1
        row += 1
    # 6.save
    # timestamp = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
    # filename = 'testdatas-' + timestamp + '.xls'
    # filename = os.path.join(topdir, 'updownload', filename)
    book.save(filename)

# please be very careful to call this function
def empty_folder_filesonly(folder):
    files = os.listdir(folder)
    for file in files:
        if file == '.gitkeep':
            continue
        file = os.path.join(folder, file)
        # todo
        # if file type is folder, here will raise IsADirectoryError
        os.remove(file)

def empty_folder_all(path):
    for root, dirs, files in os.walk(path, topdown=False):
        # print('==root==', root)
        for name in files:
            # print('==file==', name)
            if name =='.gitkeep':
                continue
            os.remove(os.path.join(root, name))
        for name in dirs:
            # print('==dir==', name)
            # os.removedirs(os.path.join(root, name))
            os.rmdir(os.path.join(root, name))

def rm_pycache(path):
     for root, dirs, files in os.walk(path, topdown=False):
        for name in dirs:
            if name == '__pycache__':
                # os.removedirs(os.path.join(root, name))
                shutil.rmtree(os.path.join(root, name))


def cleanup_log():
    empty_folder_all(logfolder)

def cleanup_pycache():
    rm_pycache(appfolder)

def write_json_to_file(o_dict, filename):
    with open(filename, 'w') as f:
        json.dump(o_dict, f, indent=4)

def read_textfile_oneline(filename):
    content = open(filename).readline()
    content = content.replace('\r', '').replace('\n','')
    return content

def get_datetime_now_obj():
    return datetime.datetime.now(tz=tz.gettz('Asia/Shanghai')).replace(microsecond=0)

def get_datetime_now_str():
    return datetime.datetime.now(tz=tz.gettz('Asia/Shanghai')).replace(microsecond=0).strftime('%Y-%m-%d %H:%M:%S')

def get_localpath_from_fullurl(url):
    # http://10.30.30.101:5000/rasp/testdata?clearsearchsession=1
    # return '/' + url.split('//', 1)[1].split('/', 1)[1]
    return '/' + url.split('//', 1)[1].split('/', 1)[1].split('?', 1)[0]
